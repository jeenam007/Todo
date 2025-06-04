from django.shortcuts import render, redirect,get_object_or_404
from django.http import HttpResponse
from todoapp.forms import TodoCreateForm,ExcelUploadForm
from .models import Todos,SubTask,Course,Student,StudentCourse
from .forms import Todos,SubTaskForm,CourseForm,StudentForm,StudentCourseMultiForm,CourseExcelUploadForm
from collections import defaultdict
from django.db import connection
from django.contrib import messages
import openpyxl
from django.db import IntegrityError
from io import BytesIO
from openpyxl import Workbook

# Create your views here.\

def index(request):
    return render(request,'index.html')         
    
def create_todo(request):
        form = TodoCreateForm(request.POST or None)
        if form.is_valid():
            form.save()
            messages.success(request,"Todo created successfully.")
            return redirect('list')
        return render(request, 'createtodo.html', {'form': form})

def upload_excel(request):
    form = ExcelUploadForm(request.POST or None, request.FILES or None)

    if request.method == 'POST' and form.is_valid():
        excel_file = request.FILES['file']
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        created = 0
        valid_statuses = ['Completed', 'Not completed']
        error_rows = []

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            task_name, user, status = row[0], row[1], row[2]
            error_message = ""

            if not all([task_name, user, status]):
                error_message = "Missing required fields."
            elif status not in valid_statuses:
                error_message = f"Invalid status: '{status}'. Must be one of {valid_statuses}."

            if error_message:
                error_rows.append((idx, task_name, user, status, error_message))
                continue  # Skip saving this row

            try:
                Todos.objects.create(task_name=task_name, user=user, status=status)
                created += 1
            except IntegrityError as e:
                error_rows.append((idx, task_name, user, status, f"Database error: {str(e)}"))

        if error_rows:
            # Create a workbook to export errors
            error_wb = Workbook()
            error_ws = error_wb.active
            error_ws.title = "Errors"
            error_ws.append(['Row Number', 'Task Name', 'User', 'Status', 'Error Message'])

            for row in error_rows:
                error_ws.append(row)

            # Save workbook to a BytesIO buffer
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=todo_upload_errors.xlsx'
            output = BytesIO()
            error_wb.save(output)
            output.seek(0)
            response.write(output.getvalue())
            return response

        # No errors — show success
        messages.success(request, f'{created} todos uploaded successfully from Excel.')
        return redirect('list')

    return render(request, 'upload_excel.html', {'form': form})


# def upload_excel(request):
#     form=ExcelUploadForm(request.POST or None,request.FILES or None)

#     if request.method=='POST' and form.is_valid():
#         excel_file=request.FILES['file']
#         wb=openpyxl.load_workbook(excel_file)
#         sheet=wb.active

#         created=0
#         valid_statuses = ['Completed', 'Not completed']

#         for idx,row in enumerate(sheet.iter_rows(min_row=2,values_only=True),start=2):
#             task_name=row[0]
#             user=row[1]
#             status=row[2]

#             if not all([task_name,user,status]):
#                 messages.error(request,f"Row{idx} is missing required fields in row{row}. Pls fill all the required fields.")
#                 return redirect('upload_todo')
            
#             # Validate status
#             if status not in valid_statuses:
#                 messages.error(request,f"Invalid status in row{idx}:'{status}'.Allowed values are:{valid_statuses}.")
            
#             try:
#             # if task_name and user and status:
#                 Todos.objects.create(task_name=task_name,user=user,status=status)
#                 created+=1
#             except IntegrityError as e:
#                 messages.error(request,f"database error on row{idx}:{str(e)}")
#                 return redirect('list')

#         messages.success(request,f'{created} todos uploaded successfully from excel.')
#         return redirect('list')
#     return render(request,'upload_excel.html',{'form':form})

def list_all_todos(request):   
    with connection.cursor() as cursor:
        cursor.execute("SELECT id,task_name,status,user FROM todoapp_todos")
        rows=cursor.fetchall()

    todos=[]
    for row in rows:
        todos.append({
            'id':row[0],
            'task_name':row[1],
            'status':row[2],
            'user':row[3],

        })
    return render(request, 'listalltodos.html', {'todos':todos})

# def list_all_todos(request):
#     todos = Todos.objects.all()  
#     context = {"todos": todos}
#     return render(request, 'listalltodos.html', context)
     

def edit_todo(request,pk):
     instance_edit=Todos.objects.get(pk=pk)
     
     if request.POST:
         frm=TodoCreateForm(request.POST,instance=instance_edit)

         if frm.is_valid():
          instance_edit.save()
          messages.success(request,"Todo updated successfully.")
         return redirect("list")
     
     else:  
         form=TodoCreateForm(instance=instance_edit)
     return render(request,"createtodo.html",{'form': form})



# def delete_todo(request,pk):
#    instance = Todos.objects.get(pk=pk)
#    instance.delete()
#    return redirect("list")

def delete_todo(request,pk):
   instance = Todos.objects.get(pk=pk)
   if SubTask.objects.filter(task_name=instance).exists():
       messages.warning(request,f'Cannot delete"{instance.task_name}"-It is already in use.')
       return redirect('list')
   instance.delete()
   messages.success(request,f'Todo "{instance.task_name}" was deleted successfully.')
   return redirect("list")


def create_subtask(request):
    if request.method == 'POST':
        form=SubTaskForm(request.POST)
        if form.is_valid():
            subtask=form.save(commit=False)
            subtask.save()
            form.save_m2m()
            return redirect('subtask_list')
    else:
        form=SubTaskForm()
    return render(request,'subtask.html',{'form':form})

def delete_subtask(request,pk):
   instance = SubTask.objects.get(pk=pk)
   instance.delete()
   return redirect("subtask_list")

# def edit_subtask(request,pk):
#     subtask=get_object_or_404(SubTask,pk=pk)

#     if request.method=="POST":
#         form=SubTaskForm(request.POST,instance=subtask)
#         if form.is_valid():
#           subtask = form.save(commit=False)
#           subtask.save()
#           form.save_m2m()
#           return redirect('subtask_list')  
#         else:
#             form=SubTaskForm(instance=subtask)
#         return render(request,'subtask.html',{'form':form})

def subtask_list(request):
    subtasks = SubTask.objects.prefetch_related('task_name').all()
    return render(request, 'subtask_list.html', {'subtasks': subtasks})



def course_create(request):
    form=CourseForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('course_list')
    return render(request,'course_create.html',{'form':form})


def courseupload_excel(request):
    form = CourseExcelUploadForm(request.POST or None, request.FILES or None)

    if request.method == 'POST' and form.is_valid():
        excel_file = request.FILES['file1']
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        created = 0
        failed_rows = []

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            course_name = row[0]
            course_code = row[1]

            if not all([course_name, course_code]):
                failed_rows.append(idx)
                continue

            try:
                Course.objects.create(course_name=course_name, course_code=course_code)
                created += 1
            except IntegrityError:
                failed_rows.append(idx)

        # Compose final message
        if created and failed_rows:
            messages.warning(
                request,
                f"{created} courses uploaded successfully. "
                f"Skipped rows: {', '.join(map(str, failed_rows))} due to missing/invalid data."
            )
        elif created:
            messages.success(request, f"{created} courses uploaded successfully from Excel.")
        elif failed_rows:
            messages.error(
                request,
                f"Upload failed. All rows skipped due to errors in rows: {', '.join(map(str, failed_rows))}."
            )

        return redirect('course_list')  # or 'list' if you want to redirect to list always

    return render(request, 'courseupload_excel.html', {'form': form})


def delete_student(request,pk):
   instance = Student.objects.get(pk=pk)
   if StudentCourse.objects.filter(student_name=instance).exists():
       messages.warning(request,f'Cannot delete"{instance.student_name}"-It is already in use.')
       return redirect('student_list')
   instance.delete()
   messages.success(request,f'Student "{instance.student_name}" was deleted successfully.')
   return redirect("student_list")


def course_list(request):
    courselist = Course.objects.all()
    return render(request,'course_list.html',{'courselist':courselist})

def course_edit(request,pk):
    course=get_object_or_404(Course,pk=pk)
    form=CourseForm(request.POST or None, instance=course)
    if form.is_valid():
        form.save()
        return redirect('course_list')
    return render(request,'course_create.html',{'form':form})

def student_edit(request,pk):
    student=get_object_or_404(Student,pk=pk)
    form=StudentForm(request.POST or None,instance=student)
    if form.is_valid():
        form.save()
        messages.success(request,"Student details updated successfully.")
        return redirect('student_list')
    return render(request,'createstudent_det.html',{'form':form})

def course_delete(request,pk):
    course=get_object_or_404(Course,pk=pk)

     # Check if this course is used in StudentCourse
    if StudentCourse.objects.filter(course_name=course).exists():
        messages.warning(request, f'Cannot delete "{course.course_name}" -it is already in use.')
        return redirect('course_list')
    course.delete()
    messages.success(request, f'Course "{course.course_name}" was deleted successfully.')
    return redirect('course_list')

def create_studentdet(request):
    form=StudentForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request,"Student details created successfully.")
        return redirect('student_list')
    return render(request,'createstudent_det.html',{'form':form})

def studentlist(request):
    student = Student.objects.all()  
    return render(request, 'student_list.html', {'student':student})

def enroll_student(request):
    form=StudentCourseMultiForm(request.POST or None)
    if form.is_valid():
        student=form.cleaned_data['student_name']
        courses=form.cleaned_data['course_name']

        enrolled_courses = []
        for course in courses:
            # Check for duplicate enrollment
            already_enrolled = StudentCourse.objects.filter(student_name=student, course_name=course).exists()
            if not already_enrolled:
                StudentCourse.objects.create(student_name=student,course_name=course)
                enrolled_courses.append(str(course))
                
            else:
                 messages.warning(request, f"{student} is already enrolled in {course}.")
        
        if enrolled_courses:
            course_list = ", ".join(enrolled_courses)
            messages.success(request, f"{student} successfully enrolled in {course_list}.")
        return redirect('enrol_list')
    else:
        form=StudentCourseMultiForm()
    return render(request,'enrol.html',{'form':form})



def student_enrol_list(request):
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT
                s.id AS student_id,
                s.student_name,
                s.qualification,
                GROUP_CONCAT(c.course_name ORDER BY c.course_name SEPARATOR ', ') AS course_names,
                MIN(sc.enrolled_on) AS first_enrolled_on
            FROM
                todoapp_studentcourse AS sc
            LEFT JOIN
                todoapp_student AS s ON sc.student_name_id = s.id
            LEFT JOIN
                todoapp_course AS c ON sc.course_name_id = c.id
            GROUP BY
                s.id, s.student_name, s.qualification
        """)
        rows = cursor.fetchall()
        grouped_enrollments = []
    for row in rows:
        enrollment = {
            'student_id': row[0],
            'student_name': row[1],
            'qualification': row[2],
            'course_names': row[3],
            'first_enrolled_on': row[4]
        }
        grouped_enrollments.append(enrollment)

        
        # columns = [col[0] for col in cursor.description]
        # grouped_enrollments = [
        #     dict(zip(columns, row))
        #     for row in cursor.fetchall()
        # ]

    return render(request, 'enrol_list.html', {
        'grouped_enrollments': grouped_enrollments
    })


def maxProfit(prices):
    min_price=float('inf')
    max_profit=0

    for price in prices:
        if price<min_price:
            min_price=price
        else:
            max_profit=max(max_profit,price-min_price)
    return max_profit

def max_profit_view(request):
    result = None
    if request.method == 'POST':
        price_str = request.POST.get('prices', '')
        try:
            prices = list(map(int, price_str.split(',')))
            result = maxProfit(prices)  # ✅ just call the function and store result
        except:
            result = "Invalid input"
    return render(request, 'profit.html', {'result': result})

